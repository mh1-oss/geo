import sys
import os

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

folder = os.path.join("dist", "a")
files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]
files.sort()

output = []

for fname in files:
    fpath = os.path.join(folder, fname)
    output.append(f"\n{'='*80}")
    output.append(f"FILE: {fname}")
    output.append(f"{'='*80}")
    wb = load_workbook(fpath, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output.append(f"\n--- Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ---")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=False):
            vals = []
            for cell in row:
                v = cell.value
                if v is not None:
                    vals.append(f"[{cell.coordinate}]{v}")
            if vals:
                output.append("  ".join(vals))
    wb.close()

with open("excel_output.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(output))

print("Done! Output saved to excel_output.txt")
